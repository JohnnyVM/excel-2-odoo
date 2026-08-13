from PyQt6.QtCore import Qt, qInfo
import openpyxl
import csv
import os

from .. import settings
from ..dependencies import get_odoo
from ..gui.model.odoomodel import OdooModel
from .fuzzyfinder import match_headers
from .textcleaner import clean_import_text


def text2many2manyfield(text, model: OdooModel):
    for data in model._data:
        if str(text) in data['display_name']:
            return [data['id']]
    return None


def text2many2onefield(text, model: OdooModel):
    for data in model._data:
        if str(text) == data['display_name']:
            return [data['id'], text]
    return None


def factoryExcelOdooModel(excel_file: str, parent):
    model = OdooModel(
        conn=get_odoo(settings.conf),
        company_id=parent.company_id,
        autoload=False)
    is_csv = os.path.splitext(excel_file)[1].lower() == '.csv'
    wb = None
    csv_file = None
    if is_csv:
        csv_file = open(excel_file, newline='', encoding='utf-8-sig')
        iter_rows = csv.reader(csv_file)
        try:
            fields = tuple(clean_import_text(value) for value in next(iter_rows))
        except StopIteration:
            csv_file.close()
            return
    else:
        wb = openpyxl.load_workbook(
            filename=excel_file, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]  # only the first
        iter_rows = sheet.iter_rows()
        fields = tuple(clean_import_text(c.value) for c in next(iter_rows))
    raw_fields = model._conn.execute_kw('product.template', 'fields_get', [], {})
    matched_fields = match_headers(fields, raw_fields)
    model._available_fields = raw_fields
    model._fields = {}
    model._column_sources = []
    model._column_selection = list(matched_fields)
    # Keep spreadsheet columns as sources; only matched Odoo fields become keys.
    for index, (header, field) in enumerate(zip(fields, matched_fields)):
        source = str(header) if str(header) not in model._fields else f"{header}_{index}"
        model._fields[source] = raw_fields[field] if field else {'string': str(header)}
        model._column_sources.append(source)
    for row in iter_rows:
        values = (
            (clean_import_text(value) for value in row)
            if is_csv
            else (clean_import_text(cell.value) for cell in row)
        )
        model._data.append(dict(zip(model._fields.keys(), values)))
    model._loadRelationalData()

    # Here is necesary transform the raw data from exel to odoo
    for column in range(model.columnCount()):
        selected = model.columnSelection()[column]
        if selected is None:
            continue
        source = model._column_sources[column]
        attributes = model._field_attributes(selected)
        for row in range(model.rowCount()):
            if attributes.get('type', None) == 'many2many':
                model._data[row][source] = text2many2manyfield(
                    model._data[row][source], model._relational_model[source])
            if attributes.get('type', None) == 'many2one':
                model._data[row][source] = text2many2onefield(
                    model._data[row][source], model._relational_model[source])

    if csv_file:
        csv_file.close()
    if wb:
        wb.close()
    return model
