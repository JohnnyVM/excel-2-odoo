import os

from PyQt6.QtWidgets import (
        QWidget,
        QFileDialog,
        QVBoxLayout,
        QDialogButtonBox)
from PyQt6.QtCore import qInfo

from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from .mainTable import TableWidget
from ..dependencies import get_odoo
from ..schema import ProductTemplate


class MainWindow(QWidget):

    def searchColumn(workbook: ReadOnlyWorksheet, column_name: str):
        header_row = map(lambda x: x.value, next(workbook.rows))
        for header_column in header_row:
            pass

    def selectExcel(self) -> str | None:
        """ TODO add support for *.xls """
        excelFile, _ = QFileDialog.getOpenFileName(
                self,
                'Select Excel file',
                os.getcwd(),
                "Excel files (*.xlsx *.xlsm)")

        if not excelFile:
            return

        return excelFile

    def loadExcel(self):
        self.mainTable.setRowCount(0)
        excelFile = self.selectExcel()
        if not excelFile:
            return

        wb = load_workbook(filename=excelFile, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]  # only the first

        iter_rows = sheet.iter_rows()
        next(iter_rows)  # discard header7
        for row in iter_rows:
            row_count = self.mainTable.rowCount()
            self.mainTable.setRowCount(row_count + 1)
            rows = [row[r] for r in range(len(self.mainTable.headerNames))]
            for column, item in enumerate(rows):
                self.mainTable.setItem(row_count, column, item.value)
            if self.mainTable.rowEmpty(row_count):
                self.mainTable.setRowCount(row_count)

        wb.close()

    def sendToOdoo(self):
        odoo = get_odoo(self.settings)
        excel_barcodes = set(bar.text() for bar in self.mainTable.columnAt(0) if bar is not None)  # column 0 is barcode
        products_ids = odoo.env['product.template'].search([('barcode', 'in', tuple(excel_barcodes))])

        products = []
        for p_id in products_ids:
            products.append(odoo.env['product.template'].browse([p_id]))

        odoo_barcodes = set(map(lambda x: x.barcode, products))

        create_barcodes = excel_barcodes - odoo_barcodes
        qInfo("{} products to create".format(len(create_barcodes)))

        productOdoo = odoo.env['product.template']
        headers = [i.text() for i in self.mainTable.headers()]
        for barcode in create_barcodes:
            for row, bar in enumerate(self.mainTable.columnAt(0)):
                if bar.text() == barcode:
                    new_product = {"type": "product", "available_in_pos": True}
                    for idx, value in enumerate(self.mainTable.rowAt(row)):
                        if value is not None:
                            new_product.update({headers[idx]: value.text()})
                    pt = ProductTemplate(**new_product)
                    p_id = productOdoo.create([pt.dict()])
                    qInfo("product({name}) created with id {id}".format(
                        name=pt.name, id=p_id))
                    pt.id = p_id

    def mainButtonsBehaviour(self, button):
        # TODO search other way different to check the text
        if button.text() == "Apply":
            self.sendToOdoo()
            return

        if button.text() == "Open":
            self.loadExcel()
            return

        raise NotImplementedError("Unknown button")

    def __init__(self, settings, *args):
        QWidget.__init__(self, *args)

        self.settings = settings
        layout = QVBoxLayout()
        self.mainTable = TableWidget()
        self.mainButtons = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Apply
                | QDialogButtonBox.StandardButton.Open)
        self.mainButtons.clicked.connect(self.mainButtonsBehaviour)
        layout.addWidget(self.mainTable)
        layout.addWidget(self.mainButtons)
        self.setLayout(layout)
