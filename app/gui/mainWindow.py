import os
from functools import partial

from PyQt6.QtWidgets import (
        QHBoxLayout,
        QFormLayout,
        QWidget,
        QFileDialog,
        QVBoxLayout,
        QDialogButtonBox)
from PyQt6.QtCore import qInfo, qDebug

from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from .mainTable import TableWidget
from ..dependencies import get_odoo
from ..schema import ProductTemplate
from .odoocombobox import OdooComboBox


class MainWindow(QWidget):

    def rowToProduct(self, row: int) -> dict:
        new_product = {"type": "product", "available_in_pos": True}
        for column in range(self.mainTable.columnCount()):
            if column < len(self.mainTable.headerNames) and self.mainTable.item(row, column):
                new_product.update({
                    self.mainTable.horizontalHeaderItem(column).text():
                    self.mainTable.item(row, column).text()
                    })

            if column in (6, 7) and self.mainTable.cellWidget(row, column):
                new_product.update({
                    self.mainTable.horizontalHeaderItem(column).text():
                    [int(self.mainTable.cellWidget(row, column).currentData())]
                    })

            if column in (8,) and self.mainTable.cellWidget(row, column):
                new_product.update({
                    self.mainTable.horizontalHeaderItem(column).text():
                    int(self.mainTable.cellWidget(row, column).currentData())
                    })
        return new_product

    def searchColumn(workbook: ReadOnlyWorksheet, column_name: str):
        header_row = map(lambda x: x.value, next(workbook.rows))
        for header_column in header_row:
            pass

    def selectExcel(self) -> str | None:
        """ TODO add support for *.xls """
        excelFile, _ = QFileDialog.getOpenFileName(
                self,
                'Select Excel file',
                os.getcwd(),
                "Excel files (*.xlsx *.xlsm)")

        if not excelFile:
            return

        return excelFile

    def loadExcel(self):
        self.mainTable.setRowCount(0)
        excelFile = self.selectExcel()
        if not excelFile:
            return

        wb = load_workbook(filename=excelFile, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]  # only the first

        iter_rows = sheet.iter_rows()
        next(iter_rows)  # discard header7
        for row in iter_rows:
            row_count = self.mainTable.rowCount()
            self.mainTable.setRowCount(row_count + 1)
            rows = [row[r] for r in range(len(self.mainTable.headerNames))]
            for column, item in enumerate(rows):
                self.mainTable.setItem(row_count, column, item.value)
            if self.mainTable.rowEmpty(row_count):
                self.mainTable.setRowCount(row_count)

        wb.close()

    def sendToOdoo(self):
        odoo = get_odoo(self.settings)
        excel_barcodes = set(bar.text() for bar in [self.mainTable.item(row, 0) for row in range(self.mainTable.rowCount())] if bar is not None)  # column 0 is barcode
        products_ids = odoo.env['product.template'].search([('barcode', 'in', tuple(excel_barcodes))])

        products = []
        for p_id in products_ids:
            products.append(odoo.env['product.template'].browse([p_id]))

        odoo_barcodes = set(map(lambda x: x.barcode, products))

        create_barcodes = excel_barcodes - odoo_barcodes
        qInfo("{} products to create".format(len(create_barcodes)))

        productOdoo = odoo.env['product.template']
        for barcode in create_barcodes:
            for row, bar in enumerate([self.mainTable.item(row, 0) for row in range(self.mainTable.rowCount())]):
                if bar.text() == barcode:
                    new_product = self.rowToProduct(row)
                    pt = ProductTemplate(**new_product)
                    p_id = productOdoo.create([pt.dict()])
                    if p_id:
                        qInfo("product({name}){values} created with id {id}".format(
                            name=pt.name, values=new_product, id=p_id))

        update_barcodes = odoo_barcodes & excel_barcodes
        for barcode in update_barcodes:
            for row, bar in enumerate([self.mainTable.item(row, 0) for row in range(self.mainTable.rowCount())]):
                if bar.text() == barcode:
                    new_product_price = self.rowToProduct(row)
                    pt = ProductTemplate(**new_product_price)
                    op = tuple(filter(lambda p: new_product_price['barcode'] == p.barcode, products))[0]
                    odoo.execute_kw('product.template', 'write', [[op.id], {'list_price': pt.list_price}])
                    qInfo("product({id}) update {values}".format(
                            id=op.id, values={'list_price': pt.list_price}))

    def mainButtonsBehaviour(self, button):
        # TODO search other way different to check the text
        if button.text() == "Apply":
            self.sendToOdoo()
            return

        if button.text() == "Open":
            self.loadExcel()
            return

        raise NotImplementedError("Unknown button")

    def set_beahviour(self):
        def clean_update(combo: OdooComboBox, company_id: int):
            combo.clear()
            combo.update(['|', ("company_id", "=", company_id), ("company_id", "=", False)])

        self.company_combobox.currentIndexChanged.connect(
                lambda c_id: partial(clean_update, self.sales_combobox)(c_id))
        self.company_combobox.currentIndexChanged.connect(
                lambda c_id: partial(clean_update, self.purchase_combobox)(c_id))

        def update_sales_column(idx: int):
            data = self.sales_combobox.currentData()
            self.mainTable.update_sales_column(data)
        self.sales_combobox.currentIndexChanged.connect(
                update_sales_column)

        def update_purchase_column(idx: int):
            data = self.purchase_combobox.currentData()
            self.mainTable.update_purchase_column(data)
        self.purchase_combobox.currentIndexChanged.connect(
                update_purchase_column)

        def update_category_column(idx: int):
            data = self.category_combobox.currentData()
            self.mainTable.update_category_column(data)
        self.category_combobox.currentIndexChanged.connect(
                update_category_column)

    def __init__(self, settings, *args):
        QWidget.__init__(self, *args)

        self.settings = settings
        layout = QVBoxLayout()
        self.mainTable = TableWidget()
        self.mainButtons = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Apply
                | QDialogButtonBox.StandardButton.Open)
        self.mainButtons.clicked.connect(self.mainButtonsBehaviour)

        formPlaceholder = QHBoxLayout()
        companyForm = QFormLayout()
        self.company_combobox = OdooComboBox('res.company')
        self.company_combobox.update()
        companyForm.addRow("company", self.company_combobox)

        tableForm = QFormLayout()
        self.sales_combobox = OdooComboBox('account.tax', domain=[("active", "=", True), ("type_tax_use", "=", "sale")])
        tableForm.addRow("Impuestos de venta", self.sales_combobox)
        self.purchase_combobox = OdooComboBox('account.tax', domain=[("active", "=", True), ("type_tax_use", "=", "purchase")])
        tableForm.addRow("Impuestos de venta", self.purchase_combobox)
        self.category_combobox = OdooComboBox('product.category')
        self.category_combobox.update()
        tableForm.addRow("Categoria", self.category_combobox)

        formPlaceholder.addLayout(companyForm)
        formPlaceholder.addLayout(tableForm)

        layout.addLayout(formPlaceholder)
        layout.addWidget(self.mainTable)
        layout.addWidget(self.mainButtons)
        self.setLayout(layout)

        self.set_beahviour()
