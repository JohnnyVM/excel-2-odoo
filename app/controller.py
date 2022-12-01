from PyQt6.QtCore import Qt
import openpyxl

from . import settings
from .dependencies import get_odoo
from .gui.model.odoomodel import OdooModel


MANDATORY_FIELDS = ['barcode', 'name', 'taxes_id', 'supplier_taxes_id', 'list_price']


def text2many2manyfield(text, model: OdooModel):
    for data in model._data:
        if str(text) in data['display_name']:
            return [data['id']]
    return None


def text2many2onefield(text, model: OdooModel):
    for data in model._data:
        if str(text) == data['display_name']:
            return [data['id'], text]
    return None


def factoryExcelOdooModel(excel_file: str, parent):
    wb = openpyxl.load_workbook(
        filename=excel_file, read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]  # only the first

    iter_rows = sheet.iter_rows()
    model = OdooModel(
        conn=get_odoo(settings.conf),
        name='Excel load',
        company_id=parent._company_id,
        autoload=False)
    fields = tuple(map(lambda c: c.value, next(iter_rows)))
    raw_fields = model._conn.execute_kw(
        'product.template',
        'fields_get',
        [fields])
    for mfield in MANDATORY_FIELDS:
        if mfield not in fields:
            raise ValueError(f"missing field {mfield}")

    # ensure the order
    for field in fields:
        if field in raw_fields:
            model._fields[field] = raw_fields[field]
        else:
            model._fields[field] = {'string': field}
    for row in iter_rows:
        model._data.append(dict(zip(model._fields.keys(), map(lambda r: r.value, row))))
    model._loadRelationalData()

    # Here is necesary transform the raw data from exel to odoo
    for column in range(model.columnCount()):
        field, attributes = tuple(model.headerData(
            column, Qt.Orientation.Horizontal, Qt.ItemDataRole.UserRole).items())[0]
        for row in range(model.rowCount()):
            if attributes.get('type', None) == 'many2many':
                model._data[row][field] = text2many2manyfield(
                    model._data[row][field], model._relational_model[field])
            if attributes.get('type', None) == 'many2one':
                model._data[row][field] = text2many2manyfield(
                    model._data[row][field], model._relational_model[field])

    wb.close()
    return model
