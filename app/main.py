#!/bin/env python
"""
    Docstring
"""

# Only needed for access to command line arguments
import os

from configparser import ConfigParser
from typing import Any

import odoorpc

from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from settings import Settings

from PyQt6.QtWidgets import (
        QApplication,
        QWidget,
        QTableWidget,
        QVBoxLayout,
        QFileDialog,
        QTableWidgetItem,
        QDialogButtonBox)
from PyQt6.QtCore import qInfo

__version__ = "0.0.1"

DEFAULT_CONFIG_FILE='config.ini'


class TableWidget(QTableWidget):
    headerNames = ("barcode", "default_code", "name", "list_price", "standard_price")
    def __init__(self, *args):
        QTableWidget.__init__(self, *args)
        self.setColumnCount(len(self.headerNames))
        self.setHorizontalHeaderLabels(self.headerNames)

    def setItem(self, row: int, column: int, item: Any):
        if item is None:
            return
        super().setItem(row, column, QTableWidgetItem(str(item)))

    def rowEmpty(self, row: int):
        for i in range(self.columnCount()):
            if self.item(row, i) is not None:
                return False
        return True

    def rowAt(self, row: int):
        return [self.item(row, column) for column in range(self.columnCount())]

    def columnAt(self, column: int):
        return [self.item(row, column) for row in range(self.rowCount())]

    def headers(self):
        return [self.horizontalHeaderItem(idx) for idx in range(len(self.headerNames))]


class Window(QWidget):
    _odoo = None

    def searchColumn(workbook: ReadOnlyWorksheet, column_name: str):
        header_row = map(lambda x :x.value, next(workbook.rows))
        for header_column in header_row:
            pass

    def loadExcel(self):
        """
            TODO add support for *.xls
        """
        self.mainTable.setRowCount(0)
        excelFile, _ = QFileDialog.getOpenFileName(
                self,
                'Select Excel file',
                os.getcwd(),
                "Excel files (*.xlsx *.xlsm)")

        if not excelFile:
            return

        wb = load_workbook(filename=excelFile, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]  # only the first

        iter_rows = sheet.iter_rows()
        next(iter_rows)  # discard header
        for row in iter_rows:
            row_count = self.mainTable.rowCount()
            self.mainTable.setRowCount(row_count + 1)
            for column, item in enumerate(row):
                self.mainTable.setItem(row_count, column, item.value)
            if self.mainTable.rowEmpty(row_count):
                self.mainTable.setRowCount(row_count)

        wb.close()

    def sendToOdoo(self):
        if not self._odoo:
            # Prepare the connection to the server
            odoo = odoorpc.ODOO(self.settings['odoo']['host'], port=self.settings['odoo']['port'])
            odoo.login(self.settings['odoo']['database'], self.settings['odoo']['user'], self.settings['odoo']['password'])
            self._odoo = odoo

        excel_barcodes = set(bar.text() for bar in self.mainTable.columnAt(0) if bar is not None)  # column 0 is barcode
        products_ids = self._odoo.env['product.template'].search([('barcode', 'in', tuple(excel_barcodes))])

        products = []
        for p_id in products_ids:
            products.append(self._odoo.env['product.template'].browse([p_id]))

        odoo_barcodes = set(map(lambda x: x.barcode, products))

        create_barcodes = excel_barcodes - odoo_barcodes
        qInfo("{} products to create".format(len(create_barcodes)))

        productOdoo = self._odoo.env['product.template']
        headers = [i.text() for i in self.mainTable.headers()]
        for barcode in create_barcodes:
            for row, bar in enumerate(self.mainTable.columnAt(0)):
                if bar.text() == barcode:
                    new_product = {"type": "product"}
                    for idx, value in enumerate(self.mainTable.rowAt(row)):
                        if value is not None:
                            new_product.update({headers[idx]: value.text()})
                    p_id = productOdoo.create([new_product])
                    qInfo("product({name}) created with id {id}".format(name=new_product['name'], id=p_id))


    def mainButtonsBehaviour(self, button):
        # TODO search other way different to check the text
        if button.text() == "Apply":
            self.sendToOdoo()
            return

        if button.text() == "Open":
            self.loadExcel()
            return

        raise NotImplementedError("Unknown button")

    def __init__(self, settings: Settings, *args):
        QWidget.__init__(self, *args)

        self.settings = settings
        layout = QVBoxLayout()
        self.mainTable = TableWidget()
        self.mainButtons = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Apply
                | QDialogButtonBox.StandardButton.Open)
        self.mainButtons.button(QDialogButtonBox.StandardButton.Apply).setEnabled(True)
        self.mainButtons.clicked.connect(self.mainButtonsBehaviour)
        layout.addWidget(self.mainTable)
        layout.addWidget(self.mainButtons)
        self.setLayout(layout)


# You need one (and only one) QApplication instance per application.
# Pass in sys.argv to allow command line arguments for your app.
# If you know you won't use command line arguments QApplication([]) works too.
app = QApplication([])
app.setApplicationVersion(__version__)
app.setApplicationName("excel-2-odoo")

if __name__=="__main__":
    config = ConfigParser()
    config.read(DEFAULT_CONFIG_FILE)

    settings = Settings(config=config)

    # Create a Qt widget, which will be our window.
    window = Window(settings=settings)

    window.show()  # IMPORTANT!!!!! Windows are hidden by default.

    # Start the event loop.
    app.exec()
